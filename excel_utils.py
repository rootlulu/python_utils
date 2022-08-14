#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""
@File    :   excel_util.py
@Time    :   2022/08/04 15:19:23
@Author  :   Ysm
@Contact :   rootlulu@163.com
"""

import logging as log
import os
import shutil
import traceback
import typing as t

from openpyxl import Workbook as WB
from openpyxl import load_workbook
from openpyxl.cell import Cell
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.styles import Alignment
from openpyxl.utils import column_index_from_string
from openpyxl.utils.exceptions import InvalidFileException

SUPPORTED_FILE_TYPE = ".xlsx"


class _WorkBook:
    """the singleton workbook if its filename is the same.

    Raises:
        InvalidFileException: the invalid file type.

    Returns:
        [WorkBook]:
    """

    # todo： multiprocess in problem.
    singleton = {}
    SUFFIX = SUPPORTED_FILE_TYPE

    def __new__(cls, filename: t.AnyStr) -> WB:
        if filename not in cls.singleton:
            if not filename.endswith(cls.SUFFIX):
                raise InvalidFileException(
                    f"The file: {filename} is invalid. the"
                    "filename must ends with .xlsx, bitch."
                )
            if os.path.exists(filename):
                log.info(f"Open the existed file: {filename}")
                wb = load_workbook(filename)
            else:
                log.info(f"Open a new file: {filename}")
                wb = WB()
            wb.filename = filename
            cls.singleton[filename] = wb
        return cls.singleton[filename]


class Validator:
    @staticmethod
    def validate(worksheet):
        # todo, the validating rules is not confirmed.
        pass


class _WorkSheetMixin:
    @staticmethod
    def _gen_col_name() -> t.AnyStr:
        """generate the col_name from A to Z.

        Yields:
            [str]: A B C ... X Y Z"
        """
        _start = 65
        _end = 91

        for i in range(_start, _end):
            yield chr(i)
        raise ValueError("The col nums is too big, which supported only A-Z!")


class Styler:
    def __init__(self, style: t.Union[dict, list]):
        if not isinstance(
            style,
            (
                dict,
                list,
            ),
        ):
            raise TypeError(
                f"the style must be a dict or list, but it's {type(style)}"
            )
        self.style = style
        self._validate()

    def __call__(self, ins: str):
        raise NotImplementedError("The method __call__ must be implemented!")

    def _validate(self):
        raise NotImplementedError("The method validate must be implemented!")


class ColStyler(Styler):
    SUPPORTED_STYLES = {
        "width": int,
    }

    def __init__(self, style: dict):
        if not isinstance(style, dict):
            raise TypeError("the col type must be a dict")
        super().__init__(style)

    def __call__(self, col: str) -> None:
        for k, v in self.style.items():
            setattr(col, k, v)

    def _validate(self):
        "validate the supported style."
        if isinstance(self.style, dict):
            for k, v in self.style.items():
                if not isinstance(v, self.SUPPORTED_STYLES[k]):
                    raise ValueError(
                        f"the style: {self.style} is not supported!"
                    )
        else:
            raise TypeError("The col style must be dict")


class RowStyler(Styler):
    """
    to be padded in the future. Not in use in current.
    """


class CellStyler(Styler):
    """
    The utils class:
        to set the style to a Cell and return it.

    Supported styles:
        `name`:         Font.name
        `style`:        Font.style
        `color`:        Font.color
        `size`:         Font.size
        `bold`:         Font.bold
        `underline`:    Font.underline
        `italic`:       Font.italic
        `vertAlign`:    Font.vertAlign
        `outline`:      Font.outline
        `shadow`:       Font.shadow

        `fgColor`:      PatternFill.fgColor
        `bgColor`:      PatternFill.bgColor
        `fill_type`:    PatternFill.fill_type

        `warpText`:     Alignment

        `width`:        width
        `height`:       height
    """

    SUPPORTED_STYLES = {
        "font": (
            "name",
            "style",
            "color",
            "size",
            "bold",
            "underline",
            "italic",
            "vertAlign",
            "outline",
            "shadow",
        ),
        "pattern_fill": ("fgColor", "bgColor", "fill_type"),
        "alignment": ("warpText",),
        # ...
        # ...
        "others": ("width", "height"),
    }

    def __call__(self, cell: Cell) -> None:

        if isinstance(self.style, dict):
            style = self.style
            self._styled_cell(style, cell)
        if isinstance(self.style, list):
            try:
                # !!! in the first index, whose col_idx is 1 but the index is 0.
                style = self.style[cell.col_idx - 1]
            except IndexError:
                style = None
            self._styled_cell(style, cell)

    def _styled_cell(self, style: dict, cell: Cell) -> None:
        if style:
            self._set_font(cell, style)
            self._set_pattern_fill(cell, style)
            # ...
            # ...
            self._set_others(cell, style)
        return cell

    def _set_font(self, cell: Cell, style: dict) -> None:
        font = Font(
            **{
                k: v
                for k, v in style.items()
                if k in self.SUPPORTED_STYLES["font"]
            }
        )
        cell.font = font

    def _set_pattern_fill(self, cell: Cell, style: dict) -> None:
        pattern_fill = PatternFill(
            **{
                k: v
                for k, v in style.items()
                if k in self.SUPPORTED_STYLES["pattern_fill"]
            }
        )
        cell.fill = pattern_fill

    def _set_aligment(self, cell: Cell, style: dict) -> None:
        alignment = Alignment(
            **{
                k: v
                for k, v in style.items()
                if k in self.SUPPORTED_STYLES["alignment"]
            }
        )
        cell.alignment = alignment

    def _set_others(self, cell: Cell, style: dict) -> None:
        for k, v in style.items():
            if k in self.SUPPORTED_STYLES["others"]:
                setattr(cell, k, v)

    def _validate(self) -> None:
        "validate the supported style."
        supported_styles = sum(self.SUPPORTED_STYLES.values(), ())
        if isinstance(self.style, dict):
            keys = self.style.keys()
        if isinstance(self.style, list):
            keys = set(sum([list(d.keys()) for d in self.style], []))
        else:
            return
        for k in keys:
            if k not in supported_styles:
                raise ValueError(
                    f"The type keyword: {k} is not supproted, the"
                    f"supported keyword is {supported_styles}"
                )


class Style:
    """
    `type_`:
            cell:
            col:
            row:

    """

    def __init__(self, style: t.Union[dict, list], type_: str = "cell"):
        if not style:
            self.styler = lambda x: x
        if type_ == "cell":
            self.styler = CellStyler(style)
        if type_ == "col":
            self.styler = ColStyler(style)
        if type_ == "row":
            self.styler = RowStyler(style)

    def __call__(self, ins):
        """
        Args:
            ins ([Col, Cell, Col]):
        """
        self.styler(ins)


class Col:
    """
    Args:
        self (Col):
        col (str):
        style (dict): supported type see ColStyler doc
        ws (WorkSheet):
    """

    def __init__(self: "Col", col: str, style: dict, ws: "WorkSheet"):
        self.ws = ws.ws
        self.col = self.ws.column_dimensions[col]
        self.style = Style(style, type_="col")

    def set(self):
        self.style(self.col)


class Row:
    """
    Args:
        self (Row):
        row_idx (int):
        style (dict): supported type see RowStyler doc
        ws (WorkSheet):
    """

    def __init__(self: "Row", row_idx: int, style: dict, ws: "WorkSheet"):
        self.ws = ws.ws
        self.row = self.ws.row_dimensions[row_idx]
        self.style = Style(style, type_="row")

    def set(self):
        self.style(self.row)


class Cell_:
    """[summary]

    Args:
        self (Cell_):
        row_idx (int):
        col_idx (int):
        style (dict): supported type see CellStyler doc.
        ws (WorkSheet):
        value ([type], optional):  Defaults to None.
    """

    def __init__(
        self: "Cell_",
        row_idx: int,
        col_idx: int,
        style: dict,
        ws: "WorkSheet",
        value=None,
    ):
        self.ws = ws.ws
        self.cell = self.ws.cell(row=row_idx, column=col_idx, value=value)
        self.style = Style(style)

    def set(self):
        self.style(self.cell)


class IterStyledCell(t.List):
    """
    The styled cell generator, yield the styled cell.
    return the raw data cells if there is not the style param or return the
        styled data otherwise.

    supported parameters see the `Style class`
    """

    def __new__(
        cls: "IterStyledCell",
        data: t.Union[list, dict],
        style: t.Union[list, dict],
        ws: "WorkSheet",
    ) -> "IterStyledCell":
        if not style:
            return data
        else:
            return super().__new__(cls)

    def __init__(
        self: "IterStyledCell",
        data: t.Union[list, dict],
        style: t.Union[list, dict],
        ws: "WorkSheet",
    ):
        self.row_data = data
        # self.style = Style(style)
        if isinstance(style, dict):
            self.style = lambda i: style
        elif isinstance(style, list):
            self.style = lambda i: style[i - 1] if len(style) >= i else {}
        self.ws = ws.ws
        self._cells = self._styled_cells()

    def __iter__(self):
        for cell in self._cells:
            yield cell.cell

    def _styled_cells(self):
        if isinstance(
            self.row_data,
            (
                list,
                tuple,
                range,
            ),
        ):
            yield from self._seq_styled_cells()
        elif isinstance(self.row_data, dict):
            yield from self._map_styled_cells()
        else:
            raise TypeError(
                f"the {type(self.row_data)} of {self.row_data} is unsupported."
            )

    def _seq_styled_cells(self):
        row_idx = self.ws._current_row + 1
        for col_idx, col_val in enumerate(self.row_data, 1):
            cell = Cell_(row_idx, col_idx, self.style(col_idx), self, col_val)
            cell.set()
            yield cell

    def _map_styled_cells(self):
        row_idx = self.ws._current_row + 1
        for col_idx, col_val in self.row_data.items():
            if isinstance(col_idx, str):
                col_idx = column_index_from_string(col_idx)
            cell = Cell_(row_idx, col_idx, self.style(col_idx), self, col_val)
            cell.set()
            yield cell


class WorkSheet(_WorkSheetMixin):
    """if there is the worksheet in the workbook, return it.
       if there is not the worksheet, return a new worksheet.

       the headers is a row that return dict with it as the keys and write to
       excel in dict format with it as the keys.

    Properties:
        `parent`: the workbook instance.

        `ws`: the worksheet instance.

        `headers`: the headers and col_names correspondency in current worksheet.
            {"name": "A", "sex": "B", "age": "C"}

    Args:
        `filename`: the destination file.

        `title`: the sheet title, default is Sheet1.

        `index`: the sheet index, default is 1. if the index is set, the title in invalid.

        `headers_idx`: the headers index in all rows. if there is a headers_idx, the to_dict
                read would start from the headers_idx row.

        `template`: if there is a template, copy it to filename and write to the copy.

        `before_styled`: the style before reading or writing the worksheet.
                        {"A": {"color" :"#0FF00"}}

        `before_styled`: the style after reading or writing but before saving the worksheet.

    """

    CLOSE = True
    SUFFIX = SUPPORTED_FILE_TYPE

    def __init__(
        self,
        filename: str,
        title: t.Optional[str] = "Sheet1",
        index: t.Optional[int] = None,
        headers_idx: t.Optional[int] = 1,
        template: t.Optional[str] = None,
        before_styled: t.Optional[dict] = None,
        after_styled: t.Optional[dict] = None,
    ):
        if template:
            if not os.path.exists(template):
                raise FileNotFoundError(
                    f"The template file: {template} not found!"
                )
            if not template.endswith(self.SUFFIX):
                raise InvalidFileException(
                    f"The file suffix must be {self.SUFFIX}!"
                )
            if template == filename:
                raise InvalidFileException(
                    f"The template file: {template} is the same"
                    f"file as destination file {filename}"
                )
            if os.path.exists(filename):
                raise FileExistsError(
                    f"The destination {filename} is existed, can't override it by tempate file."
                )
            shutil.copyfile(template, filename)

        if not isinstance(headers_idx, int):
            raise TypeError("The headers_idx must be a integer.")
        if headers_idx < 1:
            raise ValueError("The headers_idx must lager than 1.")
        self.headers_idx = headers_idx

        # the headers is None at first, initialized when read or write.
        self.headers = None

        self.before_styled = before_styled or {}
        self.after_styled = after_styled or {}
        assert isinstance(self.before_styled, dict)
        assert isinstance(self.after_styled, dict)

        self._parent_factory = _WorkBook
        parent = self._parent_factory(filename)

        if index or title:
            if index:
                try:
                    self.ws = parent.worksheets[index - 1]
                    log.info(f"open a active sheet, its index: {index-1}.")
                except IndexError as e:
                    raise ValueError(
                        "The index: {index} worksheet is not existed."
                    ) from e
            else:
                for sheet in parent.worksheets:
                    if sheet.title == title:
                        self.ws = parent[title]
                        log.info(f"open the active the sheet: {title}.")
                else:
                    if not hasattr(self, "ws"):
                        self.ws = parent.create_sheet(title)
                        log.info(f"create and open a new sheet: {title}.")
            self._gen_headers()

    def _styled_hook(self, style):
        for k, v in style.items():
            if isinstance(k, int) and k < 65535:
                self.set_row_style(k, v)
            elif (
                isinstance(k, str)
                and len(k) == 1
                and ord(k) > 64
                and ord(k) < 91
            ):
                self.set_col_style(k, v)
            else:
                raise TypeError(
                    "To set a row style in int type keys small than 65535 like: 1, 2, "
                    "the col is in str type keys in range A-Z like :A, B"
                )

    def __enter__(self) -> "WorkSheet":
        self._styled_hook(self.before_styled)
        return self

    def __exit__(self, exc_type, exc_value, traceback):
        if exc_type:
            self.close(save=False)
            log.warning(
                f"Error: {exc_type}, msg: {exc_value}, traceback: {traceback}"
            )
        else:
            self._styled_hook(self.after_styled)
            self.close()

    def _gen_headers(self, max_col: t.Optional[int] = None):
        """
        like this:
            {"name": "A", "sex": "B", "age": "C", ...}
        """
        col_names = self._gen_col_name()
        for i, row in enumerate(
            self.ws.iter_rows(max_col=max_col, values_only=True), start=1
        ):
            if i == self.headers_idx:
                self.headers = {col: next(col_names) for col in row}
                col_names.close()
                return

    def close(self, save=True):
        """
        save and close the workbook, and drop the singleton.
        """
        try:
            if save:
                self.ws.parent.save(self.ws.parent.filename)
            self.ws.parent.close()
        except:
            log.error(traceback.format_exc())
        finally:
            self._parent_factory.singleton.pop(self.ws.parent.filename)

    def validate(self):
        pass

    def to_dict(
        self,
        max_col: t.Optional[None] = None,
        show_col_names: bool = False,
        col_mapping: t.Optional[t.Mapping] = None,
    ) -> t.Generator:
        """yield the content as dict.
           if the excel like this:

                A       B       C       ...
                name     sex     age     ...
                xiaomi   female  1       ...
                lulu     male    2       ...

           if the show_col_names paramter is True:

                yield:
                    {"A": "name", "B": "sex", "C": "age"},
                    {"A": "xiaomi", "B": "female", "C": "1"},
                    {"A": "xiaomi", "B": "female", "C": "2"},
           else:
                yield:
                    {"name": "xiaomi", "sex": "female", "age": "1"},
                    {"name": "xiaomi", "sex": "female", "age": "2"},

        Args:
            `max_col`: how many lines would to read. return whole col if there
                is not a value.

            `show_col_names` (bool, optional): Defaults to False. see on the above.

            `col_mapping` (Union[bool, Dict], optional):
                for example:
                    the previous data: {"name": "男孩", "sex": "女孩", "age": "李韬"}
                    while the re_col = {"name": "名字", "sex": "性别"}
                    the after data: {"名字": "男孩", "性别": "女孩", "age": "李韬"}
                important: only function in the show_col_names is False.

        Yields:
            [Dict]: ...
        """
        if max_col:
            if not isinstance(max_col, int):
                raise TypeError("The max_col must be a integer.")
            if max_col < 1 or max_col > 26:
                raise ValueError(
                    "The max_col range must in 1 ~ 26, corresponding to"
                    "the alpha A ~ Z"
                )
        if not col_mapping:
            col_mapping = {}
        if show_col_names and col_mapping:
            log.warning(
                f"If show_col_names is set to True, the col_mapping param would be ignored."
            )
            col_mapping = {}
        for i, row in enumerate(
            self.ws.iter_rows(
                min_row=self.headers_idx, max_col=max_col, values_only=True
            )
        ):
            if not any(row):
                # jump the blank line
                continue
            if not self.headers:
                # generagte the self.headers. and write the headers to current sheet.
                self._gen_headers(max_col=max_col)
                if show_col_names:
                    yield {
                        v: k for k, v in list(self.headers.items())[:max_col]
                    }
                continue
            if show_col_names:
                yield {
                    k: v
                    for k, v in list(zip(self.headers.values(), row))[:max_col]
                }
            else:
                # drop the first row, which's the headers.
                if i == 0:
                    continue
                yield {
                    col_mapping.get(k, k): v
                    for k, v in list(zip(self.headers, row))[:max_col]
                }

    def append(
        self,
        iterable: t.Union[
            dict,
            list,
            tuple,
        ],
        style=None,
    ):
        """append a row to the worksheet.
            if the iterable is a iterable:
                if current_row is 0(means the first row) or there is not a header
                    in self, it will be the headers.
                if the current_row is not 0, it will append to the next row.
            if the iterable is a dict:
                if the current_row is 0 or there is not a header in self, it will
                    be the headers
                else append the iterable to the next row.

            IMPORTANT:
            support for dict append. if there is a headers already, the append
            will insert the value whose keys included in self.headers only.


        Args:
            `iterable` (Union[List, Dict, Tuple, range]):

            `style` (Optional[Union[Dict, List]]):
                Suppose that the row like this:
                    {"name": "xiaomi", "B": "female", "C": "1"}
                1. if the style is None:
                    return the raw row.
                2. if the style is dict: {"color": "00FF00FF"}:
                    the all cells in the row would be red style.
                3. if the style is list whose lenght is equal to the row:
                    such as: [  {"color": "00FF00FF"},
                                {"color": "00FF00FF"},
                                {"color": "00FF00FF"},]:
                    the style and the cell will be corresponding in the sequence.
                4. if the style is list but which's longer or shorter than the row's length:
                    the style adn the cell will be corresponding in the sequence.
                    if there is not the style in some cells, jump it.

        Raises:
            `ValueError`: ...
            `TypeError`: if the iterable's type is not in list, tuple and dict.
        """
        if isinstance(
            iterable,
            (
                list,
                tuple,
            ),
        ):
            if not self.headers:
                self._gen_headers()
            self.ws.append(IterStyledCell(iterable, style, self))
        elif isinstance(iterable, dict):
            if not self.headers:
                col_names = self._gen_col_name()
                self.headers = {k: next(col_names) for k in iterable}
                self.ws.append(
                    IterStyledCell(
                        [header for header in self.headers], style, self
                    )
                )
            self.ws.append(
                IterStyledCell(
                    {v: iterable.get(k) for k, v in self.headers.items()},
                    style,
                    self,
                )
            )
        else:
            raise TypeError(
                f"There is a unsupported type: {type(iterable)} for data: {iterable}."
                f"The data must in tuple, list or dict."
            )

    def set_row_style(self, row_idx: int, style: dict) -> None:
        row = Row(row_idx, style, self)
        row.set()

    def set_col_style(self, col_name: str, style: dict) -> None:
        col = Col(col_name, style, self)
        col.set()

    def set_cell(
        self, row_idx: int, col_idx: int, style: dict, value=None
    ) -> None:
        cell = Cell_(row_idx, col_idx, style, self, value)
        cell.set()


if __name__ == "__main__":
    with WorkSheet("test_bbb.xlsx", title="Sheet8") as ws:
        ws.append(
            "name",
        )
