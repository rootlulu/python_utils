#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""
@File    :   test_excel_utils.py
@Time    :   2022/08/04 09:48:44
@Author  :   Ysm
@Contact :   rootlulu@163.com
"""

import copy
import os
import shutil
from collections import namedtuple
from hashlib import md5

import pytest
from openpyxl.utils.exceptions import InvalidFileException

CUR_DIR = os.curdir


from excel_utils import WorkSheet
from excel_utils import _WorkSheetMixin

# todo What is lack of the excepted scene in Testcases. append in the future.
# todo What is lack of the excepted scene in Testcases. append in the future.
# todo What is lack of the excepted scene in Testcases. append in the future.


class ExcelInit:
    """the init file is test.xlsx. and there is a sheet and the below data:
    Sheet1:      A          B           C
                name(str)   sex(str)    age(str)
                xiaomi      female      19
                lulu        male        20
    """

    sheet = namedtuple(
        "Sheet",
        (
            "name",  # the sheet name, which is Sheet1
            "col_mapping",  # col_mapping: like {"name": "姓名"}
            "values",  # the values without headers
            "values_with_headers",  # like the name means.
            "values_with_col_mapping",  # like the name means, corresponding to the col_mapping.
        ),
    )

    FILE_READ = FILE_EXISTED = os.path.join(CUR_DIR, "./test.xlsx")
    FILE_WRITE = os.path.join(CUR_DIR, "./test_new.xlsx")
    SHEET1 = sheet(
        "Sheet1",
        {
            "name": "姓名",
            "sex": "性别",
            "age": "年龄",
        },
        [
            {
                "name": "xiaomi",
                "sex": "female",
                "age": "19",  # notice: the str.
            },
            {
                "name": "lulu",
                "sex": "male",
                "age": 20,  # notice: the number.
            },
            {
                "name": "wenbao",
                "sex": "female",
                "age": 3,  # notice: the number.
            },
        ],
        [
            {
                "A": "name",
                "B": "sex",
                "C": "age",
            },
            {
                "A": "xiaomi",
                "B": "female",
                "C": "19",  # notice: the str.
            },
            {
                "A": "lulu",
                "B": "male",
                "C": 20,  # notice:  the number.
            },
            {
                "A": "wenbao",
                "B": "female",
                "C": 3,  # notice:  the number.
            },
        ],
        [
            {
                "姓名": "xiaomi",
                "性别": "female",
                "年龄": "19",  # notice: the str.
            },
            {
                "姓名": "lulu",
                "性别": "male",
                "年龄": 20,  # notice: the number.
            },
            {
                "姓名": "wenbao",
                "性别": "female",
                "年龄": 3,  # notice: the number.
            },
        ],
    )
    SHEET2 = sheet("Sheet2", {}, [], [], [])


class TestExcelRead(ExcelInit):
    """Test read excel, for excel test.xlsx and sheet1.
        There is a little possibility to raise a Error.

    Args:
        ExcelInit ([type]): [description]
    """

    def test_with_index_and_title(self):
        # if there are the index and title, the index worked.
        with WorkSheet(self.FILE_READ, self.SHEET2.name, 1) as ws:
            for i, res in enumerate(ws.to_dict()):
                assert res == self.SHEET1.values[i]

    def test_without_col_name(self):
        # Don't display the col_name
        with WorkSheet(self.FILE_READ, self.SHEET1.name) as ws:
            for i, res in enumerate(ws.to_dict()):
                assert res == self.SHEET1.values[i]

    def test_without_col_name_with_index(self):
        # Don't display the col_name
        with WorkSheet(self.FILE_READ, index=1) as ws:
            for i, res in enumerate(ws.to_dict()):
                assert res == self.SHEET1.values[i]

    def test_with_col(self):
        # Display the col_name
        with WorkSheet(self.FILE_READ, self.SHEET1.name) as ws:
            for i, res in enumerate(ws.to_dict(show_col_names=True)):
                assert res == self.SHEET1.values_with_headers[i]

    def test_with_col_with_index(self):
        # Display the col_name
        with WorkSheet(self.FILE_READ, index=1) as ws:
            for i, res in enumerate(ws.to_dict(show_col_names=True)):
                assert res == self.SHEET1.values_with_headers[i]

    def test_with_col_mapping(self):
        with WorkSheet(self.FILE_READ, self.SHEET1.name) as ws:
            for i, res in enumerate(
                ws.to_dict(col_mapping=self.SHEET1.col_mapping)
            ):
                assert res == self.SHEET1.values_with_col_mapping[i]

    def test_with_col_mapping_with_index(self):
        with WorkSheet(self.FILE_READ, index=1) as ws:
            for i, res in enumerate(
                ws.to_dict(col_mapping=self.SHEET1.col_mapping)
            ):
                assert res == self.SHEET1.values_with_col_mapping[i]

    def test_with_col_mapping_more(self):
        """
        the col_mapping's items more than the col_name
        """
        with WorkSheet(self.FILE_READ, self.SHEET1.name) as ws:
            more_than_col_mapping = self.SHEET1.col_mapping.copy()
            more_than_col_mapping["x"] = "Y"
            for i, res in enumerate(
                ws.to_dict(col_mapping=more_than_col_mapping)
            ):
                assert res == self.SHEET1.values_with_col_mapping[i]

    def test_with_col_mapping_more_index(self):
        """
        the col_mapping's items more than the col_name
        """
        with WorkSheet(self.FILE_READ, index=1) as ws:
            more_than_col_mapping = self.SHEET1.col_mapping.copy()
            more_than_col_mapping["x"] = "Y"
            for i, res in enumerate(
                ws.to_dict(col_mapping=more_than_col_mapping)
            ):
                assert res == self.SHEET1.values_with_col_mapping[i]

    def test_with_col_mapping_little(self):
        """
        the col_mapping's items little than the col_name
        """
        with WorkSheet(self.FILE_READ, self.SHEET1.name) as ws:
            little_col_mapping = self.SHEET1.col_mapping.copy()
            little_col_mapping.pop("name")
            for i, res in enumerate(
                ws.to_dict(col_mapping=little_col_mapping)
            ):
                expected = {
                    k.replace("姓名", "name"): v
                    for k, v in self.SHEET1.values_with_col_mapping[i].items()
                }
                assert res == expected

    def test_with_col_mapping_little_index(self):
        """
        the col_mapping's items little than the col_name
        """
        with WorkSheet(self.FILE_READ, index=1) as ws:
            little_col_mapping = self.SHEET1.col_mapping.copy()
            little_col_mapping.pop("name")
            for i, res in enumerate(
                ws.to_dict(col_mapping=little_col_mapping)
            ):
                expected = {
                    k.replace("姓名", "name"): v
                    for k, v in self.SHEET1.values_with_col_mapping[i].items()
                }
                assert res == expected

    def test_with_not_integer_max_col(self):
        with WorkSheet(self.FILE_READ, self.SHEET1.name) as ws:
            with pytest.raises(TypeError):
                next(ws.to_dict(max_col="int"))

    def test_with_a_beyond_range_max_col(self):
        with WorkSheet(self.FILE_READ, self.SHEET1.name) as ws:
            with pytest.raises(ValueError):
                next(ws.to_dict(max_col=29))

    def test_with_max_col(self):
        max_col = 2
        expected = [
            {k: v for k, v in list(d.items())[:max_col]}
            for d in self.SHEET1.values
        ]
        with WorkSheet(self.FILE_READ, self.SHEET1.name, 1) as ws:
            for i, res in enumerate(ws.to_dict(max_col=max_col)):
                assert res == expected[i]

    def test_with_headers_idx(self):
        expected = []
        keys = None
        for d in self.SHEET1.values:
            if not keys:
                keys = list(d.values())
                continue
            expected.append({k: v for k, v in zip(keys, d.values())})
        print(expected)
        with WorkSheet(
            self.FILE_READ, self.SHEET1.name, 1, headers_idx=2
        ) as ws:
            for i, res in enumerate(ws.to_dict()):
                assert res == expected[i]

    def test_gen_headers_with_headers_idx_1(self):
        headers_idx = 1
        keys = self.SHEET1.values[headers_idx - 1].keys()
        expected_headers = {
            k: v for k, v in zip(keys, _WorkSheetMixin._gen_col_name())
        }
        with WorkSheet(
            self.FILE_READ, self.SHEET1.name, 1, headers_idx=headers_idx
        ) as ws:
            assert ws.headers == expected_headers

    def test_gen_headers_with_headers__idx_2(self):
        headers_idx = 2
        keys = self.SHEET1.values[headers_idx - 2].values()
        expected_headers = {
            k: v for k, v in zip(keys, _WorkSheetMixin._gen_col_name())
        }
        with WorkSheet(
            self.FILE_READ, self.SHEET1.name, 1, headers_idx=headers_idx
        ) as ws:
            assert ws.headers == expected_headers

    def test_gen_headers_without_headers_idx(self):
        keys = self.SHEET1.values[0].keys()
        expected_headers = {
            k: v for k, v in zip(keys, _WorkSheetMixin._gen_col_name())
        }
        with WorkSheet(self.FILE_READ, self.SHEET1.name, 1) as ws:
            assert ws.headers == expected_headers

    def test_to_dict_regen_headers(self):
        headers_idx = 2
        max_col = 2
        keys = self.SHEET1.values[headers_idx - 2].values()
        expected_headers = {
            k: v for k, v in list(zip(keys, _WorkSheetMixin._gen_col_name()))
        }
        with WorkSheet(
            self.FILE_READ, self.SHEET1.name, 1, headers_idx=headers_idx
        ) as ws:
            next(ws.to_dict(max_col=max_col))
            assert ws.headers == expected_headers

    def test_value_with_empty_file(self):
        with WorkSheet(self.FILE_READ, self.SHEET2.name) as ws:
            with pytest.raises(StopIteration):
                next(ws.to_dict())


class TestExcelWrite(ExcelInit):
    """
    test write excel test_new.xlsx.
    """

    def teardown_method(self):
        """remove the writing test file."""
        if os.path.exists(self.FILE_WRITE):
            os.remove(self.FILE_WRITE)

    setup_method = teardown_method

    # the below is in function scene.
    def test_dict_to_existed_sheet(self):
        """write to the existed file and existed sheet."""
        to_append = {"name": "shaobo", "sex": "male", "age": 17}
        expected = self.SHEET1.values + [to_append]
        shutil.copy(self.FILE_EXISTED, self.FILE_WRITE)
        with WorkSheet(self.FILE_WRITE, self.SHEET1.name) as ws:
            ws.append(to_append)
        with WorkSheet(self.FILE_WRITE, self.SHEET1.name) as ws:
            for i, res in enumerate(ws.to_dict()):
                assert res == expected[i]

    def test_seq_to_existed_sheet(self):
        """write to the existed file and existed sheet."""
        to_append = ["shaobo", "male", 17]
        expected = self.SHEET1.values + [
            {"name": "shaobo", "sex": "male", "age": 17}
        ]
        shutil.copy(self.FILE_EXISTED, self.FILE_WRITE)
        with WorkSheet(self.FILE_WRITE, self.SHEET1.name) as ws:
            ws.append(to_append)
        with WorkSheet(self.FILE_WRITE, self.SHEET1.name) as ws:
            for i, res in enumerate(ws.to_dict()):
                assert res == expected[i]

    def test_dict_to_new_sheet(self):
        """write to the new sheet in the existed file."""
        to_append = expected = {"name": "shaobo", "sex": "male", "age": 17}
        shutil.copy(self.FILE_EXISTED, self.FILE_WRITE)
        with WorkSheet(self.FILE_WRITE, self.SHEET2.name) as ws:
            ws.append(to_append)
        with WorkSheet(self.FILE_WRITE, self.SHEET2.name) as ws:
            for _, res in enumerate(ws.to_dict()):
                assert res == expected

    def test_seq_to_new_sheet(self):
        """write to the new sheet in the existed file.
        there is not any content in ws.to_dict() because the read is no need.
        """
        to_append = ["name", "sex", "age"]
        shutil.copy(self.FILE_EXISTED, self.FILE_WRITE)
        with WorkSheet(self.FILE_WRITE, self.SHEET2.name) as ws:
            ws.append(to_append)
        with pytest.raises(StopIteration):
            with WorkSheet(self.FILE_WRITE, self.SHEET2.name) as ws:
                next(ws.to_dict())

    def test_dict_to_new_excel(self):
        """write to a new excel, and its sheet must be new also."""
        to_append = expected = {"name": "shaobo", "sex": "male", "age": 17}
        with WorkSheet(self.FILE_WRITE, self.SHEET2.name) as ws:
            ws.append(to_append)
        with WorkSheet(self.FILE_WRITE, self.SHEET2.name) as ws:
            for i, res in enumerate(ws.to_dict()):
                assert res == expected

    def test_seq_to_new_excel(self):
        """write to a new excel, and its sheet must be new also."""
        to_append1 = ["name", "sex", "age"]
        to_append2 = ["shaobo", "male", 17]
        expected = {"name": "shaobo", "sex": "male", "age": 17}
        with WorkSheet(self.FILE_WRITE, self.SHEET2.name) as ws:
            ws.append(to_append1)
            ws.append(to_append2)
        with WorkSheet(self.FILE_WRITE, self.SHEET2.name) as ws:
            for _, res in enumerate(ws.to_dict()):
                assert res == expected

    def test_no_create_new_file_when_crash(self):
        """
        Don't save the new file if what raise a exception.
        """
        to_append_string = "name: shaobo, sex: male, age: 17"
        with pytest.raises(TypeError):
            with WorkSheet(self.FILE_WRITE, self.SHEET2.name) as ws:
                ws.append(to_append_string)
        assert os.path.exists(self.FILE_WRITE) is False

    def test_no_change_existed_when_crash(self):
        """
        Don't save the original file if what raise a exception.
        """
        shutil.copy(self.FILE_EXISTED, self.FILE_WRITE)
        to_append_string = "name: shaobo, sex: male, age: 17"
        with pytest.raises(TypeError):
            with WorkSheet(self.FILE_WRITE, self.SHEET2.name) as ws:
                ws.append(to_append_string)
        assert os.path.exists(self.FILE_WRITE) is True
        with WorkSheet(self.FILE_WRITE, self.SHEET2.name) as ws:
            for i, res in enumerate(ws.to_dict()):
                assert res == self.SHEET1.values[i]

    def test_template(self):
        # todo the both files are the same, but the testcase is not passed.
        with WorkSheet(self.FILE_WRITE, template=self.FILE_EXISTED):
            pass
        with open(self.FILE_EXISTED, "rb") as f:
            content = f.read()
            md5_value1 = md5(content).digest()
        with open(self.FILE_WRITE, "rb") as f:
            content = f.read()
            md5_value2 = md5(content).digest()
        print(md5_value1, md5_value2)
        assert md5_value1 == md5_value2

    def test_template_already_existed(self):
        with WorkSheet(self.FILE_WRITE, template=self.FILE_EXISTED):
            pass
        with pytest.raises(FileExistsError):
            with WorkSheet(self.FILE_WRITE, template=self.FILE_EXISTED):
                pass

    def test_template_same_as_file(self):
        with pytest.raises(InvalidFileException):
            with WorkSheet(self.FILE_READ, template=self.FILE_EXISTED):
                pass

    def test_write_with_headers_idx_2(self):
        headers_idx = 2
        shutil.copy(self.FILE_EXISTED, self.FILE_WRITE)
        expected = copy.deepcopy(self.SHEET1.values)
        expected.append({"name": "1", "sex": "2", "age": "3"})
        with WorkSheet(
            self.FILE_WRITE, self.SHEET1.name, 1, headers_idx=headers_idx
        ) as ws:
            ws.append({"xiaomi": "1", "female": "2", "19": "3"})
        with WorkSheet(self.FILE_WRITE, self.SHEET1.name, 1) as ws:
            for i, res in enumerate(ws.to_dict()):
                assert res == expected[i]

    # the below is abnormal scene.
    def test_string(self):
        """
        write a string to the sheet.
        if the type is not in tuple, list and dict. riase a TypeError
        """
        to_append_string = "name: shaobo, sex: male, age: 17"
        with WorkSheet(self.FILE_WRITE, self.SHEET2.name) as ws:
            with pytest.raises(TypeError):
                ws.append(to_append_string)


class TestExcelSetStyle(ExcelInit):
    def teardown_method(self):
        # remove the writing test file.
        if os.path.exists(self.FILE_WRITE):
            os.remove(self.FILE_WRITE)

    setup_method = teardown_method

    def test_write_color_with_seq_style(self):
        to_append = ["name", "sex", "age"]
        style = {"color": "00800000"}
        with WorkSheet(self.FILE_WRITE, self.SHEET2.name) as ws:
            ws.append(to_append, style=style)
        # You can validate the color while open the file.
        with WorkSheet(self.FILE_WRITE, self.SHEET2.name) as ws:
            assert (
                ws.ws["A1"].font.color.rgb
                == ws.ws["B1"].font.color.rgb
                == ws.ws["C1"].font.color.rgb
                == style["color"]
            )

    def test_write_color_with_dict_style(self):
        to_append = {"name": "shaobo", "sex": "male", "age": 16}
        style = [
            {"color": "00FF00FF"},
            {"color": "00FF00FF"},
            {"color": "00FF00FF"},
        ]
        with WorkSheet(self.FILE_WRITE, self.SHEET2.name) as ws:
            ws.append(to_append, style=style)
        # You can validate the color while open the file.
        with WorkSheet(self.FILE_WRITE, self.SHEET2.name) as ws:
            assert ws.ws["A2"].font.color.rgb == style[0]["color"]
            assert ws.ws["B2"].font.color.rgb == style[1]["color"]
            assert ws.ws["C2"].font.color.rgb == style[2]["color"]

    def test_write_size_with_dict_style(self):
        to_append = {"name": "shaobo", "sex": "male", "age": 16}
        style = {"size": 15}
        with WorkSheet(self.FILE_WRITE, self.SHEET2.name) as ws:
            ws.append(to_append, style=style)
        # You can validate the color while open the file.
        with WorkSheet(self.FILE_WRITE, self.SHEET2.name) as ws:
            assert (
                ws.ws["A2"].font.size
                == ws.ws["B2"].font.size
                == ws.ws["C2"].font.size
                == style["size"]
            )

    def test_write_size_with_seq_style(self):
        to_append = {"name": "shaobo", "sex": "male", "age": 16}
        style = [{"size": 15}, {"size": 35}, {"size": 96}]
        with WorkSheet(self.FILE_WRITE, self.SHEET2.name) as ws:
            ws.append(to_append, style=style)
        # You can validate the color while open the file.
        with WorkSheet(self.FILE_WRITE, self.SHEET2.name) as ws:
            assert ws.ws["A2"].font.size == style[0]["size"]
            assert ws.ws["B2"].font.size == style[1]["size"]
            assert ws.ws["C2"].font.size == style[2]["size"]

    def test_write_size_and_color_with_dict_style(self):
        to_append = {"name": "shaobo", "sex": "male", "age": 16}
        style = {"size": 15, "color": "00FF00FF"}
        with WorkSheet(self.FILE_WRITE, self.SHEET2.name) as ws:
            ws.append(to_append, style=style)
        # You can validate the color while open the file.
        with WorkSheet(self.FILE_WRITE, self.SHEET2.name) as ws:
            assert (
                ws.ws["A2"].font.size
                == ws.ws["B2"].font.size
                == ws.ws["C2"].font.size
                == style["size"]
            )
            assert (
                ws.ws["A2"].font.color.rgb
                == ws.ws["B2"].font.color.rgb
                == ws.ws["C2"].font.color.rgb
                == style["color"]
            )

    def test_write_size_and_color_with_seq_style(self):
        to_append = {"name": "shaobo", "sex": "male", "age": 16}
        style = [
            {"size": 15, "color": "00FF00FF"},
            {"size": 25, "color": "0033CCCC"},
            {"size": 35, "color": "00CCFFCC"},
        ]
        with WorkSheet(self.FILE_WRITE, self.SHEET2.name) as ws:
            ws.append(to_append, style=style)
        # You can validate the color while open the file.
        with WorkSheet(self.FILE_WRITE, self.SHEET2.name) as ws:
            assert ws.ws["A2"].font.size == style[0]["size"]
            assert ws.ws["B2"].font.size == style[1]["size"]
            assert ws.ws["C2"].font.size == style[2]["size"]
            assert ws.ws["A2"].font.color.rgb == style[0]["color"]
            assert ws.ws["B2"].font.color.rgb == style[1]["color"]
            assert ws.ws["C2"].font.color.rgb == style[2]["color"]

    def test_set_col_before_style(self):
        before_style = {"A": {"width": 49}}
        with WorkSheet(
            self.FILE_WRITE, self.SHEET2.name, before_styled=before_style
        ) as ws:
            ws.append({"name": "shaobo", "sex": "male", "age": 19})
        with WorkSheet(self.FILE_WRITE, self.SHEET2.name) as ws:
            assert (
                ws.ws.column_dimensions["A"].width
                == before_style["A"]["width"]
            )

    def test_set_col_after_style(self):
        after_style = {"A": {"width": 49}}
        with WorkSheet(
            self.FILE_WRITE, self.SHEET2.name, before_styled=after_style
        ) as ws:
            ws.append({"name": "shaobo", "sex": "male", "age": 19})
        with WorkSheet(self.FILE_WRITE, self.SHEET2.name) as ws:
            assert (
                ws.ws.column_dimensions["A"].width == after_style["A"]["width"]
            )

    def test_set_row_before_style(self):
        # todo, the code is not finished.
        pass

    def test_set_row_after_style(self):
        # todo, the code is not finished.
        pass

    def test_set_cell(self):
        style = {"color": "0033CCCC", "size": 50}
        initalized = {"name": "shaobo", "sex": "male", "age": 19}
        expected = {"name": "test", "sex": "male", "age": 19}
        with WorkSheet(self.FILE_WRITE, self.SHEET2.name) as ws:
            ws.append(initalized)
            ws.set_cell(2, 1, style, expected["name"])
        with WorkSheet(self.FILE_WRITE, self.SHEET2.name) as ws:
            for res in ws.to_dict():
                assert res == expected
            assert ws.ws["A2"].font.size == style["size"]
            assert ws.ws["A2"].font.color.rgb == style["color"]


class TestExcelValidator(ExcelInit):
    pass


abcd

class TestExcelMixin(_WorkSheetMixin):
    def test_gen_col_name(self):
        expected = [
            "A",
            "B",
            "C",
            "D",
            "E",
            "F",
            "G",
            "H",
            "I",
            "J",
            "K",
            "L",
            "M",
            "N",
            "O",
            "P",
            "Q",
            "R",
            "S",
            "T",
            "U",
            "V",
            "W",
            "X",
            "Y",
            "Z",
        ]
        # the A-Z will passed, but the next will raise a ValueError
        with pytest.raises(
            ValueError,
            match="The col nums is too big, which supported only A-Z!",
        ):
            for i, col_name in enumerate(self._gen_col_name()):
                assert col_name == expected[i]

defg

2222
3333
4444